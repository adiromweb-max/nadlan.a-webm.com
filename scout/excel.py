"""
בניית קובץ האקסל — RTL, כותרות קפואות, רוחב עמודות, וקישורים חיים.

חמישה גיליונות:
  1. הזדמנויות        — לפי דירוג (לבדוק דחוף / שווה בדיקה)
  2. כל המודעות       — כל השדות, כולל דגל איכות נתונים
  3. מעקב ירידות מחיר — כולל סימון ירידה חריגה כשגיאת נתונים
  4. השוואת אזורים    — חציון רשמי + עליית ערך (CAGR)
  5. מגמת אזורים      — חציון ₪/מ"ר לכל שנה, הבסיס ל-CAGR ולגרפים

בכל שורת מודעה שני קישורים אמיתיים (לחיצים): המודעה ביד2, וחיפוש
בנדל"ן ממשלתי הממוקד לעיר/אזור של אותה מודעה.
"""
import logging
import shutil
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .scoring import TIER_URGENT, TIER_WORTH

log = logging.getLogger(__name__)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
URGENT_FILL = PatternFill("solid", fgColor="C6EFCE")     # לבדוק דחוף
WORTH_FILL = PatternFill("solid", fgColor="FFF2CC")      # שווה בדיקה
SUSPECT_FILL = PatternFill("solid", fgColor="FCE4E4")    # חשוד / בדיקה ידנית
LINK_FONT = Font(color="1155CC", underline="single")

MONEY = '#,##0 "₪"'
# סימן אחוז **מילולי**. פורמט האחוז של Excel (0.0%) מכפיל ב-100 בתצוגה,
# ואז 3.7 היה מוצג כ-370%. אל תחליף.
PCT = '0.0"%"'
PCT_SIGNED = '+0.0"%";-0.0"%";0.0"%"'


def _link_cols(headers, *names):
    """
    {שם_עמודה: אינדקס} עבור עמודות הקישורים.

    נגזר מרשימת הכותרות ולא נכתב כמספר קשיח: עמודה שתתווסף באמצע הייתה
    מזיזה את האינדקסים ותולה את הקישורים על העמודה הלא נכונה בשקט.
    """
    return {n: headers.index(n) for n in names if n in headers}


def _write_sheet(ws, headers, rows, formats=None, note=None, links=None,
                 row_fill=None):
    """
    כותב גיליון מעוצב.
    links: {שם_עמודה: index_בשורה_של_ה-URL} — הופך את התא לקישור לחיץ.
    row_fill: פונקציה (index_שורה) -> PatternFill|None לצביעת שורה שלמה.
    """
    ws.sheet_view.rightToLeft = True
    start_row = 1

    if note:
        ws.cell(row=1, column=1, value=note).font = Font(italic=True, size=10,
                                                         color="666666")
        ws.merge_cells(start_row=1, start_column=1, end_row=1,
                       end_column=max(len(headers), 1))
        start_row = 3

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[start_row].height = 30

    for r, row in enumerate(rows, start_row + 1):
        fill = row_fill(r - start_row - 1) if row_fill else None
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            header = headers[c - 1]
            if formats and header in formats:
                cell.number_format = formats[header]
            if links and header in links:
                url = row[links[header]]
                if url:
                    cell.hyperlink = url
                    cell.font = LINK_FONT
            elif fill:
                cell.fill = fill

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    for c, h in enumerate(headers, 1):
        widths = [len(str(h))] + [len(str(row[c - 1])) for row in rows
                                  if row[c - 1] is not None]
        ws.column_dimensions[get_column_letter(c)].width = min(
            max(max(widths) + 3, 10), 55)

    if not rows:
        ws.cell(row=start_row + 1, column=1,
                value="— אין נתונים בגיליון הזה בריצה הנוכחית —").font = Font(
                    italic=True, color="888888")


def _opportunity_rows(scored):
    """שורות גיליון ההזדמנויות — רק מה שדורג 'לבדוק דחוף' או 'שווה בדיקה'."""
    return [s for s in scored
            if s.get("stage") == "final" and s.get("tier") in (TIER_URGENT, TIER_WORTH)]


def build_workbook(scored, drops, city_rows, run_meta, out_dir, threshold,
                   areas=None, sections=None):
    """
    scored: רשימת dicts (מודעה + ניקוד), ממוינת יורד לפי ציון.
    drops: רשימת ירידות מחיר מ-db.price_drop_history.
    city_rows: רשימת dicts להשוואת אזורים.
    areas: {area_key: area_info} — לגיליון המגמה.
    מחזיר (path_dated, path_latest).
    """
    wb = Workbook()

    # ---------- 1. הזדמנויות ----------
    ws = wb.active
    ws.title = "הזדמנויות"
    headers = ["דירוג", "סוג ההזדמנות", "עיר", "שכונה", "מחיר", "₪ למ\"ר",
               "חדרים", "גודל (מ\"ר)", "חציון ₪/מ\"ר בהשוואה", "מס' קומפים",
               "רמת ההשוואה", "פער %", "ירידת מחיר %", "ימים באוויר",
               "תשואה ברוטו %", "עליית ערך שנתית %", "ציון",
               "איכות הנתונים", "הסבר", "קישור למודעה", "נדל\"ן ממשלתי"]
    opps = _opportunity_rows(scored)
    rows = [[
        s.get("tier_he"), s.get("opportunity_type_he"), s.get("city"),
        s.get("neighborhood"), s.get("price"), _round(s.get("price_per_sqm")),
        s.get("rooms"), s.get("size_sqm"),
        _round((s.get("comp") or {}).get("comp_median_ppm")),
        (s.get("comp") or {}).get("comp_count"),
        (s.get("comp") or {}).get("comp_match_level"),
        _round(s.get("gap_pct"), 1), _round(s.get("drop_pct"), 1),
        s.get("days_on_market"), _round(s.get("yield_pct"), 1),
        _round(s.get("area_cagr_pct"), 1), s.get("score"),
        s.get("data_quality"), s.get("reason"),
        s.get("url"), s.get("nadlan_link"),
    ] for s in opps]

    def _fill_for(i):
        s = opps[i]
        if (s.get("comp") or {}).get("suspect"):
            return SUSPECT_FILL
        return URGENT_FILL if s.get("tier") == TIER_URGENT else WORTH_FILL

    _write_sheet(ws, headers, rows,
                 formats={"מחיר": MONEY, "₪ למ\"ר": MONEY,
                          "חציון ₪/מ\"ר בהשוואה": MONEY, "פער %": PCT,
                          "ירידת מחיר %": PCT, "תשואה ברוטו %": PCT,
                          "עליית ערך שנתית %": PCT_SIGNED},
                 links=_link_cols(headers, "קישור למודעה", 'נדל"ן ממשלתי'),
                 row_fill=_fill_for,
                 note=('מדורג לפי "לבדוק דחוף" ואז "שווה בדיקה". הפער מחושב מול '
                       'חציון ₪/מ"ר של עסקאות שנסגרו בפועל באזור הספציפי. '
                       'ירוק = לבדוק דחוף, צהוב = שווה בדיקה, אדום = פער חשוד '
                       'הדורש בדיקה ידנית. '
                       f'נוצר {run_meta.get("run_date")}.'))

    # ---------- 2. כל המודעות ----------
    ws2 = wb.create_sheet("כל המודעות")
    headers2 = ["עיר", "שכונה", "מחיר", "₪ למ\"ר", "חדרים", "גודל (מ\"ר)",
                "שלב הבדיקה", "דירוג", "סוג ההזדמנות",
                "חציון ₪/מ\"ר בהשוואה", "מס' קומפים", "רמת ההשוואה",
                "פער %", "ירידת מחיר %", "ימים באוויר", "תשואה ברוטו %",
                "עליית ערך שנתית %", "ציון",
                "פער מדורג %", "רמת ההשוואה (סולם)", "רמת ביטחון", "סימון ערך",
                "מצב הנכס", "בסיס ההשוואה",
                "איכות הנתונים", "הסבר", "מקור",
                "קישור למודעה", "נדל\"ן ממשלתי"]
    rows2 = [[
        s.get("city"), s.get("neighborhood"), s.get("price"),
        _round(s.get("price_per_sqm")), s.get("rooms"), s.get("size_sqm"),
        "עסקאות אמיתיות" if s.get("stage") == "final" else "סינון מקדים",
        s.get("tier_he"), s.get("opportunity_type_he"),
        _round((s.get("comp") or {}).get("comp_median_ppm")),
        (s.get("comp") or {}).get("comp_count"),
        (s.get("comp") or {}).get("comp_match_level"),
        _round(s.get("gap_pct"), 1), _round(s.get("drop_pct"), 1),
        s.get("days_on_market"), _round(s.get("yield_pct"), 1),
        _round(s.get("area_cagr_pct"), 1), s.get("score"),
        _round((s.get("value") or {}).get("value_gap_pct"), 1),
        (s.get("value") or {}).get("comp_level_he"),
        (s.get("value") or {}).get("confidence_he"),
        (s.get("value") or {}).get("value_tag"),
        s.get("condition_text"), s.get("benchmark_label"),
        s.get("data_quality"), s.get("reason"), s.get("source"),
        s.get("url"), s.get("nadlan_link"),
    ] for s in scored]
    _write_sheet(ws2, headers2, rows2,
                 formats={"מחיר": MONEY, "₪ למ\"ר": MONEY,
                          "חציון ₪/מ\"ר בהשוואה": MONEY, "פער %": PCT,
                          "ירידת מחיר %": PCT, "תשואה ברוטו %": PCT,
                          "עליית ערך שנתית %": PCT_SIGNED,
                          "פער מדורג %": PCT},
                 links=_link_cols(headers2, "קישור למודעה", 'נדל"ן ממשלתי'),
                 note='כל המודעות שנסרקו בריצה הזו. מודעות בשלב "סינון מקדים" '
                      'הושוו לחציון המבוקש של יד2 בלבד; רק מי שעברה את השער '
                      'נבדקה מול עסקאות שנסגרו בפועל. "מס\' קומפים" סופר '
                      'תצפיות רבעון של מחיר עסקה ממוצע — לא עסקאות בודדות '
                      '(המקור הרשמי אינו חושף עסקה בודדת).')

    # ---------- 3. מכ"ם ירידות מחיר ----------
    # שורה אחת לכל **מודעה** (ולא לכל אירוע שינוי): התמונה המצטברת היא
    # מה שמעניין — כמה ירדה בסך הכול, בכמה שלבים, ומה המסלול המלא.
    ws3 = wb.create_sheet("מכ\"ם ירידות מחיר")
    radar = (sections or {}).get("price_drop_radar")
    headers3 = ["עיר", "שכונה", "מחיר מקורי", "מחיר נוכחי", "ירידה מצטברת %",
                "מס' ירידות", "ירידה אחרונה %", "תאריך שינוי אחרון",
                "מסלול המחיר", "ירידה חדה", "חדרים", "גודל (מ\"ר)",
                "פער מול השוואה %", "רמת ההשוואה", "רמת ביטחון",
                "איכות הנתונים", "קישור למודעה"]
    if radar is None:
        # נפילה חזרה: אין מדורים (למשל קריאה ישירה מבדיקה) — בונים מהאירועים
        radar = [{"city": d.get("city"), "original_price": d.get("old_price"),
                  "current_price": d.get("new_price"),
                  "total_drop_pct": d.get("drop_pct"), "num_drops": 1,
                  "last_drop_pct": d.get("drop_pct"),
                  "last_change_at": d.get("changed_at"), "rooms": d.get("rooms"),
                  "size_sqm": d.get("size_sqm"), "suspect": d.get("suspect"),
                  "url": d.get("url")}
                 for d in sorted(drops or [], key=lambda x: -(x.get("drop_pct") or 0))]
    rows3 = [[
        r.get("city"), r.get("neighborhood"), r.get("original_price"),
        r.get("current_price"), _round(r.get("total_drop_pct"), 1),
        r.get("num_drops"), _round(r.get("last_drop_pct"), 1),
        r.get("last_change_at"), r.get("history_text"),
        "כן" if r.get("sharp") else "", r.get("rooms"), r.get("size_sqm"),
        _round(r.get("value_gap_pct"), 1), r.get("comp_level_he"),
        r.get("confidence_he"),
        "בדיקה ידנית — ירידה חריגה" if r.get("suspect") else "תקין",
        r.get("url"),
    ] for r in radar]
    _write_sheet(ws3, headers3, rows3,
                 formats={"מחיר מקורי": MONEY, "מחיר נוכחי": MONEY,
                          "ירידה מצטברת %": PCT, "ירידה אחרונה %": PCT,
                          "פער מול השוואה %": PCT},
                 links=_link_cols(headers3, "קישור למודעה"),
                 row_fill=lambda i: (SUSPECT_FILL if radar[i].get("suspect")
                                     else (URGENT_FILL if radar[i].get("sharp")
                                           else None)),
                 note='כל מודעה שהמחיר שלה ירד, לפי הירידה המצטברת (מבוסס '
                      'price_history ב-SQLite). "ירידה מצטברת" = (מחיר מקורי − '
                      'מחיר נוכחי) / מחיר מקורי × 100, כלומר כל הירידות יחד ולא '
                      'רק האחרונה. "מסלול המחיר" מציג את השרשרת המלאה עם תאריכים. '
                      'ירידה מעל 60% מסומנת כשגיאת נתונים — לא מנוקדת ולא נשלחת '
                      'בהתראה. ירוק = ירידה חדה.')

    # ---------- 3ב. ערך יחסי ----------
    ws3b = wb.create_sheet("ערך יחסי")
    headers3b = ["דירוג", "עיר", "שכונה", "מחיר", "₪ למ\"ר", "חדרים",
                 "גודל (מ\"ר)", "פער מתחת לחציון %", "חציון ההשוואה ₪/מ\"ר",
                 "מס' תצפיות", "רמת ההשוואה", "אזור ההשוואה", "רמת ביטחון",
                 "סימון", "ציון", "קישור למודעה", "נדל\"ן ממשלתי"]
    best = (sections or {}).get("best_relative_value") or []
    rows3b = [[
        i, r.get("city"), r.get("neighborhood"), r.get("price"),
        _round(r.get("price_per_sqm")), r.get("rooms"), r.get("size_sqm"),
        _round(r.get("value_gap_pct"), 1), _round(r.get("value_median_ppm")),
        r.get("value_count"), r.get("comp_level_he"), r.get("value_area"),
        r.get("confidence_he"), r.get("value_tag"), r.get("score"),
        r.get("url"), r.get("nadlan_link"),
    ] for i, r in enumerate(best, 1)]
    _write_sheet(ws3b, headers3b, rows3b,
                 formats={"מחיר": MONEY, "₪ למ\"ר": MONEY,
                          "חציון ההשוואה ₪/מ\"ר": MONEY,
                          "פער מתחת לחציון %": PCT},
                 links=_link_cols(headers3b, "קישור למודעה", 'נדל"ן ממשלתי'),
                 row_fill=lambda i: (SUSPECT_FILL if best[i].get("suspect")
                                     else (WORTH_FILL if best[i].get(
                                         "unverified_attractive") else URGENT_FILL)),
                 note='המודעות הרחוקות ביותר מתחת לחציון ההשוואה שלהן — '
                      '**בלי תלות בשער ההתראה המחמיר**. הפער נמדד מול הרמה '
                      'הספציפית ביותר שיש בה מספיק תצפיות: שכונה → יישוב → '
                      'יישובים דומים → מרחב. "רמת ביטחון" אומרת עד כמה ההשוואה '
                      'מקומית: גבוה = שכונה/יישוב, בינוני = קבוצת יישובים/מרחב, '
                      'נמוך = אין די תצפיות באף רמה. צהוב = אטרקטיבי לכאורה '
                      'אך לא אומת, אדום = פער חשוד.')

    # ---------- 4. השוואת אזורים ----------
    ws4 = wb.create_sheet("השוואת אזורים")
    headers4 = ["עיר", "קוד יישוב", "חציון ₪ למ\"ר (רשמי)",
                "מחיר עסקה ממוצע 12ח'", "שינוי מחירים 12ח' %",
                "עליית ערך שנתית % (CAGR)", "שנות נתונים",
                "עסקאות 12ח'", "נסרקו בריצה", "מודעות פעילות",
                "אוכלוסייה", "גרסת נתונים רשמית", "נדל\"ן ממשלתי"]
    rows4 = [[
        c.get("city"), c.get("setl_code"), _round(c.get("median_ppsqm")),
        _round(c.get("avg_price_12m")), _round(c.get("price_change_pct"), 1),
        _round(c.get("cagr_pct"), 1), c.get("years_covered"),
        c.get("deals_12m_display"), c.get("scanned_now"), c.get("active_listings"),
        c.get("population"), c.get("data_version"), c.get("nadlan_link"),
    ] for c in sorted(city_rows, key=lambda x: (x.get("median_ppsqm") or 0),
                      reverse=True)]
    _write_sheet(ws4, headers4, rows4,
                 formats={"חציון ₪ למ\"ר (רשמי)": MONEY,
                          "מחיר עסקה ממוצע 12ח'": MONEY,
                          "שינוי מחירים 12ח' %": PCT_SIGNED,
                          "עליית ערך שנתית % (CAGR)": PCT_SIGNED},
                 links=_link_cols(headers4, 'נדל"ן ממשלתי'),
                 note='נתוני נדל"ן ממשלתי (עסקאות שנסגרו בפועל). '
                      'עמודת "עסקאות 12ח\'" אינה נגישה ציבורית מהשרת הזה, ולכן '
                      'מוצגת כמספר רבעונים עם עסקאות מפורסמות.')

    # ---------- 5. מגמת אזורים ----------
    ws5 = wb.create_sheet("מגמת אזורים")
    headers5 = ["אזור", "רמה", "עיר", "שנה", "חציון ₪ למ\"ר",
                "מחיר עסקה חציוני", "רבעונים עם עסקאות",
                "עליית ערך שנתית % (CAGR)", "גרסת נתונים"]
    rows5 = []
    for area in (areas or {}).values():
        for y in (area.get("years") or []):
            rows5.append([
                area.get("area_name"),
                "שכונה" if area.get("area_level") == "neighborhood" else "יישוב",
                area.get("city"), y.get("year"), _round(y.get("median_ppm")),
                _round(y.get("median_price")), y.get("deal_quarters"),
                _round(area.get("cagr_pct"), 1), area.get("data_version"),
            ])
    rows5.sort(key=lambda r: (str(r[0]), r[3] or 0))
    _write_sheet(ws5, headers5, rows5,
                 formats={"חציון ₪ למ\"ר": MONEY, "מחיר עסקה חציוני": MONEY,
                          "עליית ערך שנתית % (CAGR)": PCT_SIGNED},
                 note='חציון ₪/מ"ר לכל שנה — הבסיס לחישוב עליית הערך (CAGR) '
                      'ולגרפי המגמה ב-out/charts. שנה עם פחות מ-4 רבעונים היא '
                      'כיסוי חלקי. הגרפים והנתונים המלאים גם ב-out/latest.json.')

    out_dir.mkdir(parents=True, exist_ok=True)
    dated = out_dir / f"nadlan_{date.today():%Y%m%d}.xlsx"
    wb.save(dated)
    latest = out_dir / "latest.xlsx"
    shutil.copyfile(dated, latest)
    log.info("נשמר אקסל: %s (וגם %s) — %d הזדמנויות, %d מודעות",
             dated.name, latest.name, len(rows), len(rows2))
    return dated, latest


def _level_he(level):
    return {"neighborhood": "שכונה", "settlement": "יישוב"}.get(level, "—")


def _round(v, nd=0):
    if v is None:
        return None
    try:
        return round(float(v), nd) if nd else int(round(float(v)))
    except (TypeError, ValueError):
        return None
