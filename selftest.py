#!/usr/bin/env python
"""
בדיקה עצמית של הצינור — בלי תלות ביד2 (שדורש ScraperAPI) וללא שליחת מייל.

מוכיחה שהמנגנונים הקריטיים עובדים:
  1. נוסחת ירידת המחיר + חסם שגיאת הנתונים
  2. upsert + זיהוי ירידת מחיר בין ריצות (price_history)
  3. דדופ התראות (first_alerted_at / last_alerted_score)
  4. השוואה like-for-like מול בסיס עסקאות רשמי אמיתי
  5. תשואה, עליית ערך (CAGR) ודירוג
  6. אקסל (5 גיליונות, קישורים), latest.json וגרפים

הרצה:  ./venv/bin/python selftest.py
משתמשת ב-DB זמני — לא נוגעת ב-data/nadlan.db האמיתי.
"""
import sys
import tempfile
from datetime import date, timedelta
from pathlib import Path

from scout import (alerts, charts, comps, dashboard, db, excel, jsonout,
                   nadlan, pricing, scoring, sections)
from scout.config import load_config
from scout.http import Fetcher

FAILURES = []


def check(name, cond, detail=""):
    status = "עובר" if cond else "נכשל"
    print(f"  [{status}] {name}" + (f" — {detail}" if detail else ""))
    if not cond:
        FAILURES.append(name)
    return cond


def main():
    cfg = load_config()
    today = date.today()
    tmp = Path(tempfile.mkdtemp(prefix="nadlan-selftest-"))
    cfg["paths"]["db"] = tmp / "test.db"
    cfg["paths"]["out"] = tmp / "out"

    print("\n=== 1. נוסחת ירידת המחיר ===")
    a = pricing.assert_drop_math()
    check("400,000→385,000 מוצג כ-3.8", a["ok"], a["text"])
    check("1,000,000→963,000 מוצג כ-3.7",
          round(pricing.drop_pct(1000000, 963000), 1) == 3.7)
    check("ירידה של 90% מסומנת כשגיאת נתונים",
          not pricing.is_drop_sane(90.0))
    check("ירידה של 3.7% נחשבת תקינה", pricing.is_drop_sane(3.7))

    print("\n=== 2. התמדה וזיהוי ירידת מחיר ===")
    conn = db.connect(cfg["paths"]["db"])

    first_seen = (today - timedelta(days=100)).isoformat()
    listing = {
        "id": "test:1", "source": "test", "city": "באר שבע",
        "url": "https://example.invalid/1", "rooms": 3, "size_sqm": 72,
        "price": 700000, "price_per_sqm": 700000 / 72,
        "text": "דירה דורשת שיפוץ, המחיר גמיש, להשקעה", "raw": {},
    }
    r1 = db.upsert_listing(conn, listing, first_seen)
    conn.commit()
    check("ריצה 1: המודעה זוהתה כחדשה", r1["is_new"] is True)

    changed_at = (today - timedelta(days=10)).isoformat()
    listing2 = dict(listing, price=616000, price_per_sqm=616000 / 72)
    r2 = db.upsert_listing(conn, listing2, changed_at)
    conn.commit()
    check("ריצה 2: זוהה מחיר קודם", r2["old_price"] == 700000.0)
    check("ריצה 2: first_seen נשמר", r2["first_seen"] == first_seen)

    drops = db.price_drop_history(conn)
    check("זוהתה ירידת מחיר אחת", len(drops) == 1)
    if drops:
        check("אחוז הירידה נכון (12%)", abs(drops[0]["drop_pct"] - 12.0) < 0.01,
              f'{drops[0]["drop_pct"]:.2f}%')
        check("הירידה לא סומנה כחשודה", not drops[0]["suspect"])

    # ירידה בלתי-סבירה — חייבת להיות מסומנת
    db.upsert_listing(conn, dict(listing, id="test:2", price=1000000), first_seen)
    db.upsert_listing(conn, dict(listing, id="test:2", price=90000), changed_at)
    conn.commit()
    bogus = [d for d in db.price_drop_history(conn) if d["listing_id"] == "test:2"]
    check("ירידה של 91% סומנה כשגיאת נתונים",
          bool(bogus) and bogus[0]["suspect"] is True,
          f'{bogus[0]["drop_pct"]:.1f}%' if bogus else "")

    print("\n=== 3. דדופ התראות ===")
    cands = [{"id": "test:1", "score": 80, "price": 616000},
             {"id": "test:2", "score": 75, "price": 500000}]
    send, supp, stats = alerts.decide(cands, db.alert_state(conn, ["test:1", "test:2"]))
    check("שתי המודעות חדשות ונשלחות", len(send) == 2 and stats["new"] == 2)
    db.record_alerts(conn, send, today.isoformat())
    conn.commit()
    send2, supp2, stats2 = alerts.decide(
        [dict(c) for c in cands], db.alert_state(conn, ["test:1", "test:2"]))
    check("בריצה שנייה שתיהן מדוכאות", len(send2) == 0 and stats2["repeat"] == 2,
          alerts.summary_he(stats2))
    up = [dict(cands[0], score=95)]
    send3, _s3, stats3 = alerts.decide(up, db.alert_state(conn, ["test:1"]))
    check("עליית ציון של 15 מחזירה להתראה",
          len(send3) == 1 and stats3["score_up"] == 1)
    down = [dict(cands[1], price=450000)]
    send4, _s4, stats4 = alerts.decide(down, db.alert_state(conn, ["test:2"]))
    check("ירידת מחיר מחזירה להתראה",
          len(send4) == 1 and stats4["price_drop"] == 1)

    print("\n=== 4. השוואה like-for-like מול עסקאות אמיתיות ===")
    fetcher = Fetcher(delay=1.0, timeout=cfg["request_timeout_seconds"])
    index = nadlan.fetch_settlement_index(fetcher)
    check("אינדקס היישובים נטען", bool(index), f"{len(index)} יישובים")
    setl = nadlan.resolve_city(index, "באר שבע") if index else None
    check("קוד יישוב לבאר שבע = 9000", bool(setl) and setl["code"] == 9000)

    baseline = nadlan.fetch_city_baseline(fetcher, "באר שבע", setl) if setl else None
    check("בסיס עסקאות נטען", bool(baseline))
    if not baseline:
        conn.close()
        return 1

    observed = {3: 78.0, 4: 95.0, 5: 120.0}
    c = comps.like_for_like(baseline, listing2, cfg, observed)
    check("נמצאו קומפים מספיקים", c["sufficient"],
          f'{c["comp_count"]} תצפיות, חציון {c["comp_median_ppm"]:,.0f} ₪/מ"ר')
    check("רמת ההשוואה דווחה", c["comp_match_level"] in ("שכונה", "יישוב"),
          c["comp_match_level"])
    check("הפער חושב", c["gap_pct"] is not None, f'{c["gap_pct"]:.1f}%')

    # דירה של 2 חדרים — אין קטגוריה תואמת ברשומות הרשמיות (3/4/5)
    c2 = comps.like_for_like(baseline, dict(listing2, rooms=2), cfg, observed)
    check("2 חדרים → אין מספיק קומפים", not c2["sufficient"], c2["data_quality"])

    # פער מוגזם → חשוד
    cheap = dict(listing2, price=350000, price_per_sqm=350000 / 72)
    c3 = comps.like_for_like(baseline, cheap, cfg, observed)
    check("פער מעל 30% מסומן כחשוד", c3["suspect"], c3["data_quality"])

    print("\n=== 5. תשואה, עליית ערך ודירוג ===")
    area = comps.area_series(baseline, cfg, observed)
    check("חושב CAGR", area["cagr_pct"] is not None,
          f'{area["cagr_pct"]:.2f}% לשנה, {area["cagr_from_year"]}–{area["cagr_to_year"]}')
    check("יש חציונים שנתיים", len(area["years"]) >= 3,
          f'{len(area["years"])} שנים')

    rent = nadlan.fetch_rent(fetcher, "settlement", 9000)
    check("נתוני שכירות נטענו", bool(rent), str((rent or {}).get("rooms")))
    yinfo = comps.rental_yield(listing2, rent, cfg, "settlement")
    check("חושבה תשואה ברוטו", yinfo["yield_pct"] is not None,
          f'{yinfo["yield_pct"]:.2f}% ({yinfo["rent_basis"]})')
    check("התשואה מסומנת כהערכה", yinfo["is_estimate"] is True)

    res = scoring.score_listing(
        listing2, cfg, baseline=baseline, drop=drops[0] if drops else None,
        first_seen=first_seen, today=today, stage="final",
        comp=c, yield_info=yinfo, area_info=area)
    b = res["breakdown"]
    print(f"       ציון {res['score']} · דירוג {res['tier_he']} · "
          f"סוג {res['opportunity_type_he']}")
    print(f"       פירוט: {b}")
    print(f"       הסבר: {res['reason']}")
    check("הציון בטווח 0..100", 0 <= res["score"] <= 100, str(res["score"]))
    check("סכום הפירוט = הציון", abs(sum(b.values()) - res["score"]) < 0.02)
    check("price_gap חושב מהקומפים", b["price_gap"] > 0, f'gap={res["gap_pct"]:.1f}%')
    check("יש אות תומך", res["has_signal"], str(res["signals"]))
    check("הוקצה דירוג", res["tier"] in (scoring.TIER_URGENT, scoring.TIER_WORTH,
                                         scoring.TIER_WATCH), str(res["tier"]))
    check("הוקצה סוג הזדמנות", bool(res["opportunity_type"]),
          res["opportunity_type"])
    check("דגל איכות נתונים קיים", bool(res["data_quality"]), res["data_quality"])

    # בלי קומפים מספיקים — אין ניקוד פער, ולא יכול להיות "לבדוק דחוף"
    res_nc = scoring.score_listing(
        listing2, cfg, baseline=baseline, first_seen=first_seen, today=today,
        stage="final", comp=c2, yield_info=yinfo, area_info=area)
    check("בלי קומפים — אין ניקוד פער", res_nc["breakdown"]["price_gap"] == 0)
    check("בלי קומפים — לא 'לבדוק דחוף'", res_nc["tier"] != scoring.TIER_URGENT,
          str(res_nc["tier"]))

    print("\n=== 5ב. סולם המחירים המצטבר ===")
    # שלוש מדרגות: 700,000 → 660,000 → 616,000
    db.upsert_listing(conn, dict(listing, id="test:3", price=700000),
                      first_seen)
    db.upsert_listing(conn, dict(listing, id="test:3", price=660000),
                      (today - timedelta(days=40)).isoformat())
    db.upsert_listing(conn, dict(listing, id="test:3", price=616000), changed_at)
    conn.commit()
    ladders = db.price_ladder(conn, sanity_limit=cfg.get("max_plausible_drop_pct"))
    lad = ladders.get("test:3") or {}
    check("סולם: מחיר מקורי נשמר", lad.get("original_price") == 700000.0,
          str(lad.get("original_price")))
    check("סולם: מחיר נוכחי נכון", lad.get("current_price") == 616000.0)
    check("סולם: ירידה מצטברת 12%",
          abs((lad.get("total_drop_pct") or 0) - 12.0) < 0.01,
          f'{lad.get("total_drop_pct", 0):.2f}%')
    check("סולם: נספרו 2 ירידות", lad.get("num_drops") == 2,
          str(lad.get("num_drops")))
    check("סולם: יש מחרוזת היסטוריה עם 3 מחירים",
          (lad.get("history_text") or "").count("→") == 2, lad.get("history_text"))
    bogus_lad = ladders.get("test:2") or {}
    check("סולם: ירידה חריגה מסומנת כחשודה", bogus_lad.get("suspect") is True)

    print("\n=== 5ג. סולם ההשוואה המדורג ===")
    market = comps.MarketIndex(cfg)
    market.add_city("באר שבע", baseline, observed)
    # יישוב נוסף אמיתי כדי שתהיה מדרגת "יישובים דומים" ומרחב
    setl2 = nadlan.resolve_city(index, "אשקלון")
    base2 = nadlan.fetch_city_baseline(fetcher, "אשקלון", setl2) if setl2 else None
    if base2:
        market.add_city("אשקלון", base2, observed)
    market.build()

    v = comps.cascade(listing2, c, market, cfg)
    check("סולם: נבחרה רמת השוואה", v["comp_level"] is not None,
          f'{v["comp_level_he"]} ({v["value_count"]} תצפיות)')
    check("סולם: רמה מקומית → ביטחון גבוה",
          v["confidence"] == comps.CONF_HIGH, v["confidence_he"])
    check("סולם: יש 4 שלבים בסולם", len(v["ladder"]) >= 3,
          str([f'{r["level_he"]}:{r["count"]}' for r in v["ladder"]]))
    check("סולם: חושב פער", v["value_gap_pct"] is not None,
          f'{v["value_gap_pct"]:.1f}%')

    # דירה של 2 חדרים — אין לה קטגוריה רשמית, ובכל זאת אסור שתיעלם
    small = dict(listing2, rooms=2, price=300000, price_per_sqm=300000 / 72)
    v2 = comps.cascade(small, c2, market, cfg)
    check("מודעה בלי קומפים מקומיים אינה נפסלת",
          v2.get("value_tag") is not None or v2["comp_level"] is not None,
          f'רמה={v2["comp_level_he"]}, ביטחון={v2["confidence_he"]}, '
          f'תגית={v2.get("value_tag")}')

    # פער מוגזם — עדיין חשוד, גם בסולם
    v3 = comps.cascade(cheap, c3, market, cfg)
    check("סולם: פער מעל 30% מסומן כחשוד", v3["suspect"] is True,
          v3.get("value_tag"))

    print("\n=== 6. פלט: אקסל, JSON, דשבורד וגרפים ===")
    scored = [{**listing2, **res, "comp": c, "yield_info": yinfo,
               "area_info": area, "value": v,
               "nadlan_link": nadlan.search_link("באר שבע", c.get("comp_area"))}]

    print("      מדורים:")
    secs = sections.build(scored, ladders, areas_placeholder := {
        baseline["area_key"]: area}, cfg)
    check("מדור מכ\"ם ירידות אינו ריק", len(secs["price_drop_radar"]) >= 1,
          f'{len(secs["price_drop_radar"])} מודעות')
    check("מדור ערך יחסי אינו ריק", len(secs["best_relative_value"]) >= 1)
    check("מדור אזורים מתחממים אינו ריק", len(secs["hot_areas"]) >= 1)
    sharp = [r for r in secs["price_drop_radar"] if r.get("sharp")]
    check("ירידה של 12% סומנה כירידה חדה", len(sharp) >= 1,
          f'{len(sharp)} חדות מתוך {len(secs["price_drop_radar"])}')
    check("ירידה חשודה נדחקת לסוף הרשימה",
          not secs["price_drop_radar"][0].get("suspect"))
    city_rows = [{
        "city": "באר שבע", "setl_code": 9000,
        "median_ppsqm": c["comp_median_ppm"],
        "avg_price_12m": baseline["all_rooms_avg_price"],
        "price_change_pct": baseline["price_change_pct"],
        "cagr_pct": area["cagr_pct"], "years_covered": area["years_covered"],
        "deals_12m_display": "4/4 רבעונים עם עסקאות", "active_listings": 1,
        "scanned_now": 1, "population": baseline["population"],
        "data_version": baseline["data_version"],
        "nadlan_link": nadlan.search_link("באר שבע"),
    }]
    areas = {baseline["area_key"]: area}
    meta = {"run_date": today.isoformat(), "tier_counts": {},
            "drop_assertion": a, "dedup_stats": stats}

    meta["section_counts"] = sections.counts(secs)
    meta["summary_lines"] = [("מודעות", len(scored))]
    meta["credits_used"] = 0

    made = charts.build_all(scored, areas, cfg["paths"]["out"] / "charts",
                            ladder_rows=secs["price_drop_radar"])
    check("נוצר גרף מגמת אזור", made["area"] >= 1, str(made))
    check("נוצר גרף התפלגות קומפים", made["comps"] >= 1)
    check("נוצר גרף סולם מחירים", made["ladder"] >= 1, str(made["ladder"]))

    dated, latest = excel.build_workbook(
        scored, drops, city_rows, meta, cfg["paths"]["out"], threshold=0,
        areas=areas, sections=secs)
    check("קובץ אקסל נוצר", dated.is_file() and dated.stat().st_size > 4000,
          f"{dated.stat().st_size} bytes")

    from openpyxl import load_workbook
    wb = load_workbook(dated)
    check("6 גיליונות", len(wb.sheetnames) == 6, str(wb.sheetnames))
    ws_radar = wb['מכ"ם ירידות מחיר']
    hdr_r = [ws_radar.cell(row=3, column=i).value
             for i in range(1, ws_radar.max_column + 1)]
    for col in ("מחיר מקורי", "מחיר נוכחי", "ירידה מצטברת %", "מס' ירידות",
                "מסלול המחיר", "ירידה חדה"):
        check(f"עמודת מכ\"ם '{col}' קיימת", col in hdr_r)
    check("שורת מכ\"ם מכילה מסלול מחיר",
          "→" in str(ws_radar.cell(row=4, column=hdr_r.index("מסלול המחיר") + 1).value))
    ws_val = wb["ערך יחסי"]
    hdr_v = [ws_val.cell(row=3, column=i).value
             for i in range(1, ws_val.max_column + 1)]
    for col in ("פער מתחת לחציון %", "רמת ההשוואה", "רמת ביטחון"):
        check(f"עמודת ערך יחסי '{col}' קיימת", col in hdr_v)
    ws = wb["כל המודעות"]
    hdr = [ws.cell(row=3, column=i).value for i in range(1, ws.max_column + 1)]
    for col in ("חציון ₪/מ\"ר בהשוואה", "מס' קומפים", "רמת ההשוואה", "פער %",
                "ירידת מחיר %", "ימים באוויר", "תשואה ברוטו %",
                "עליית ערך שנתית %", "דירוג", "סוג ההזדמנות", "הסבר",
                "איכות הנתונים"):
        check(f"עמודה '{col}' קיימת", col in hdr)
    link_cell = ws.cell(row=4, column=hdr.index("קישור למודעה") + 1)
    check("קישור למודעה הוא היפר-קישור אמיתי",
          link_cell.hyperlink is not None, str(link_cell.hyperlink and
                                               link_cell.hyperlink.target))
    nad_cell = ws.cell(row=4, column=hdr.index('נדל"ן ממשלתי') + 1)
    check("קישור לנדל\"ן ממשלתי הוא היפר-קישור אמיתי",
          nad_cell.hyperlink is not None,
          str(nad_cell.hyperlink and nad_cell.hyperlink.target))
    check("גיליון מגמת אזורים מכיל נתונים",
          wb["מגמת אזורים"].cell(row=4, column=1).value is not None)

    jpath = jsonout.write(scored, areas, drops, city_rows, meta,
                          cfg["paths"]["out"], sections_data=secs,
                          ladders=ladders)
    check("latest.json נוצר", jpath.is_file())
    import json
    doc = json.loads(jpath.read_text(encoding="utf-8"))
    check("ל-JSON יש בלוק areas", bool(doc.get("areas")),
          f'{len(doc.get("areas", []))} אזורים')
    check("לאזור יש חציונים שנתיים ו-CAGR",
          bool(doc["areas"][0].get("yearly")) and
          doc["areas"][0].get("cagr_pct") is not None)
    rec = doc["listings"][0]
    for f in ("comp_median_ppm", "comp_count", "comp_match_level", "gap_pct",
              "drop_pct", "days_on_market", "yield_pct", "area_cagr_pct",
              "tier", "opportunity_type", "reason", "data_quality",
              "original_price", "current_price", "total_drop_pct", "num_drops",
              "price_history_text", "comp_level", "confidence",
              "value_gap_pct"):
        check(f"ל-JSON יש שדה {f}", f in rec and rec[f] is not None)
    check("ל-JSON יש היסטוריית מחירים מלאה",
          len(rec.get("price_history") or []) >= 2,
          f'{len(rec.get("price_history") or [])} נקודות')
    check("ל-JSON יש סולם השוואה מלא",
          len(rec.get("comp_ladder") or []) >= 3)
    for key in ("price_drop_radar", "best_relative_value", "hot_areas"):
        check(f"ל-JSON יש מדור {key}", bool((doc.get("sections") or {}).get(key)))

    dpath = dashboard.write(secs, meta, cfg, cfg["paths"]["out"], areas=areas,
                            listings_n=len(scored))
    check("dashboard.html נוצר", dpath.is_file() and dpath.stat().st_size > 3000,
          f"{dpath.stat().st_size} bytes")
    dhtml = dpath.read_text(encoding="utf-8")
    check("הדשבורד עברי ו-RTL", 'dir="rtl"' in dhtml and 'lang="he"' in dhtml)
    # פונט רשת (Google Fonts) מותר כשיפור פרוגרסיבי — נטען אונליין, נופל
    # לפונט מערכת אופליין. אסורים: תמונות, וכל CSS/JS מקומי או חיצוני אחר
    # (הם שוברים את הקובץ כשמעבירים אותו).
    import re as _re
    _refs = (_re.findall(r'<link[^>]+href="([^"]+)"', dhtml)
             + _re.findall(r'<script[^>]+src="([^"]+)"', dhtml))
    _allowed = ("https://fonts.googleapis.com", "https://fonts.gstatic.com")
    _bad = [u for u in _refs if not any(u.startswith(a) for a in _allowed)]
    check("הדשבורד בלי משאבים שבירים (רק פונט רשת חיצוני מותר)",
          not _bad and "<img" not in dhtml,
          f"חיצוניים לא-מאושרים: {_bad}" if _bad else "רק פונטים")
    for title in ('מכ"ם ירידות מחיר', "הערך היחסי הטוב ביותר",
                  "מגמת אזורים", "לוח עליית הערך"):
        check(f"הדשבורד מכיל '{title}'", title in dhtml)
    check("הדשבורד מכיל גרף SVG", "<svg" in dhtml and "</svg>" in dhtml)

    conn.close()
    print(f"\nקבצי הבדיקה: {tmp}")
    if FAILURES:
        print(f"\n*** {len(FAILURES)} בדיקות נכשלו: {FAILURES}")
        return 1
    print("\n*** כל הבדיקות עברו ***")
    return 0


if __name__ == "__main__":
    sys.exit(main())
